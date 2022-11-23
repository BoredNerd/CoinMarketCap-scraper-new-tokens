from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.series import DataPoint
import subprocess
import os
import datetime
import requests

wb = Workbook()
ws = wb.active
ws.title = "Token Analysis"
ws.append(["Name", "MarketCap"])
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.sheet_properties.tabColor = "1072BA"
ws.sheet_view.zoomScale = 130  # set the zoom level

html_text = requests.get("https://coinmarketcap.com/new/").text  # request the page HTML
soup = BeautifulSoup(html_text, "lxml") # parse the page
token = soup.select("tbody")

count = 1 #To enumerate the tokens printed in terminal
for tr in token:
    for td in tr:
        for name_token in td.contents[2].select('p[font-weight="semibold"]'): # table data of Token's names
            for mkcap in td.contents[6]: # table data of Token's Market Cap
                print(f"{count}) {name_token.text} Market Cap : {mkcap.text}")
                count += 1
                if mkcap.text != "--" and int(mkcap.text.replace("$", "").replace(",", "")) < 1000_000:
                    ws.append([name_token.text, int(mkcap.text.replace(",", "").replace("$", ""))])

for row in ws.iter_rows(min_row=1, max_col=2, max_row=ws.max_row):
    for x in row:
        if x.row == 1:  # first row
            x.font = Font(bold=True)
        if x.column == 2 and x.row != 1:  # column B and row>1
            ws[x.coordinate].number_format = '#,##0.00$'  # sets number format so the chart can be created correctly based on these values
        x.alignment = Alignment(horizontal='center', vertical="center")  # Set Alignment of every cell
        x.font = Font(size = 14)

#Create table
tab = Table(displayName="Table1", ref=f"A1:B{ws.max_row}") # Starts from A1 to B{max_row}
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

chart = PieChart()
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
chart.add_data(data)
chart.set_categories(categories)
chart.title = "Token Analysis"
chart.height = 20
chart.width = 28
ws.add_chart(chart, "E2")
# Cut the maximum token's Market Cap slice out of the pie
for col in ws.iter_cols(min_row=2, min_col=2, max_row=ws.max_row, values_only=True):
    idx_maxMkap = col.index(max(col))
slice = DataPoint(idx=idx_maxMkap, explosion=10)
chart.series[0].data_points = [slice]

if not os.path.exists("token_history"):  # if the folder "token_history" does not exist, then it creates the folder
    os.makedirs("token_history")
namefile = "token_history/NewToken_" + datetime.datetime.now().strftime("%b-%d-%Y_%H-%M") + ".xlsx"
wb.save(namefile)  # Save the Excel file in the folder "token_history"
openExcel = ((os.getcwd()) + "\\" + namefile).replace("\\", "\\\\").replace("/", "\\\\")
subprocess.Popen(openExcel, shell="false")  # open automatically the Excel file
